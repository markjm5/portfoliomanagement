import sys
import stat
import requests
import time
import os.path
import pandas as pd
import xml.etree.ElementTree as ET
import openpyxl
import zipfile
import json
import copy
import re
import yfinance as yf
import investpy as invest
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from datetime import date
from datetime import datetime as dt
from dateutil import relativedelta
from bs4 import BeautifulSoup

############################
# Data Retrieval Functions #
############################

def get_page(url):
  # When website blocks your request, simulate browser request: https://stackoverflow.com/questions/56506210/web-scraping-with-python-problem-with-beautifulsoup
  header={'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2227.0 Safari/537.36'}
  page = requests.get(url=url,headers=header)

  try:
      page.raise_for_status()
  except requests.exceptions.HTTPError as e:
      # Whoops it wasn't a 200
      raise Exception("Http Response (%s) Is Not 200: %s" % (url, str(page.status_code)))

  return page

def get_page_selenium(url):

  #Selenium Browser Emulation Tool
  chrome_options = Options()
  chrome_options.add_argument("--headless")
  chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows Phone 10.0; Android 4.2.1; Microsoft; Lumia 640 XL LTE) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.0.2311.135 Mobile Safari/537.36 Edge/12.10166")

  driver = webdriver.Chrome(ChromeDriverManager().install(),options=chrome_options)
  driver.get(url)
  driver.implicitly_wait(10)  
  time.sleep(5)
  html = driver.page_source
  driver.close()

  return html


def get_stlouisfed_data(series_code):
  url = "https://api.stlouisfed.org/fred/series/observations?series_id=%s&api_key=8067a107f45ff78491c1e3117245a0a3&file_type=json" % (series_code,)

  resp = get_page(url)

  json = resp.json() 
  
  df = pd.DataFrame(columns=["DATE",series_code])
  # Convert the Date into Time Series Date
  # https://www.youtube.com/watch?v=UFuo7EHI8zc

  for i in range(len(json["observations"])):
    if(json["observations"][i]["value"] == '.'):
      obs = '0.00'
    else:
      obs = json["observations"][i]["value"]
    df = df.append({"DATE": json["observations"][i]["date"], series_code: obs}, ignore_index=True)

  df['DATE'] = pd.to_datetime(df['DATE'],format='%Y-%m-%d')
  df[series_code] = df[series_code].astype('float64') 
  #df = df.set_index('DATE')

  print("Retrieved Data for Series %s" % (series_code,))

  return df


# Get US GDP from St Louis FRED #
"""
def get_us_gdp_fred():
  df_GDPC1 = get_stlouisfed_data('GDPC1')

  df_GDPC1['GDPQoQ'] = (df_GDPC1['GDPC1'] - df_GDPC1['GDPC1'].shift()) / df_GDPC1['GDPC1'].shift()
  df_GDPC1['GDPYoY'] = (df_GDPC1['GDPC1'] - df_GDPC1['GDPC1'].shift(periods=4)) / df_GDPC1['GDPC1'].shift(periods=4)

  #Calculate QoQ Annualized growth rate =((1 + df_GDPC1['GDPQoQ'])^4)-1
  df_GDPC1['GDPQoQ_ANNUALIZED'] = ((1 + df_GDPC1['GDPQoQ']) ** 4) - 1

  return df_GDPC1


# Get GDP Data from St Louis FRED #
def get_gdp_fred(series_name):

  df_GDP = get_stlouisfed_data(series_name)

  df_GDP['GDPQoQ'] = (df_GDP[series_name] - df_GDP[series_name].shift()) / df_GDP[series_name].shift()
  df_GDP['GDPYoY'] = (df_GDP[series_name] - df_GDP[series_name].shift(periods=4)) / df_GDP[series_name].shift(periods=4)

  #Calculate QoQ Annualized growth rate =((1 + df_GDPC1['GDPQoQ'])^4)-1
  df_GDP['GDPQoQ_ANNUALIZED'] = ((1 + df_GDP['GDPQoQ']) ** 4) - 1

  return df_GDP
"""

def get_data_fred(series_name, col_name, period):

  df = get_stlouisfed_data(series_name)
  df = df.rename(columns={series_name: col_name})

  if(period=='Q'):
    df['%s_QoQ' % (col_name,)] = (df[col_name] - df[col_name].shift()) / df[col_name].shift()
    df['%s_YoY' % (col_name,)] = (df[col_name] - df[col_name].shift(periods=4)) / df[col_name].shift(periods=4)
    df['%s_QoQ_ANNUALIZED' % (col_name,)] = ((1 + df['%s_QoQ' % (col_name)]) ** 4) - 1

  elif(period=='M'):
    df['%s_MoM' % (col_name,)] = (df[col_name] - df[col_name].shift()) / df[col_name].shift()
    df['%s_QoQ' % (col_name,)] = (df[col_name] - df[col_name].shift(periods=4)) / df[col_name].shift(periods=4)
    df['%s_YoY' % (col_name,)] = (df[col_name] - df[col_name].shift(periods=12)) / df[col_name].shift(periods=12)
    df['%s_QoQ_ANNUALIZED' % (col_name,)] = ((1 + df['%s_QoQ' % (col_name)]) ** 12) - 1

  return df

def get_oecd_data(dataset, dimensions, params):

  dim_args = ['+'.join(d) for d in dimensions]
  dim_str = '.'.join(dim_args).replace('..','.')

  date_range = dimensions[3][0]
  match date_range:
    case 'Q':
      date_range = 'QTR'
    case 'M':
      date_range = 'MTH'

  url = "https://stats.oecd.org/restsdmx/sdmx.ashx/GetData/%s/%s/all?startTime=%s&endTime=%s" % (dataset, dim_str,params['startTime'],params['endTime'])
  #file_path = '/Users/markmukherjee/Documents/PythonProjects/PortfolioManagement/XML/%s' % params['filename'] 
  file_path = "%s/XML/%s" % (sys.path[0],params['filename'])

  resp = requests.get(url=url)

  try:
      resp.raise_for_status()
  except requests.exceptions.HTTPError as e:
      # Whoops it wasn't a 200

      if(resp.status_code == 400):
        #It didnt work with the original order of the params so lets try again
        url = "https://stats.oecd.org/restsdmx/sdmx.ashx/GetData/%s/%s.%s.%s.%s/all?startTime=%s&endTime=%s" % (dataset, dim_args[1],dim_args[0],dim_args[2],dim_args[3],params['startTime'],params['endTime'])

        #clean up any situation where the format of the url is broken
        url = url.replace('..','.')

        resp = get_page(url)

      else:
            # Whoops it wasn't a 200 
            raise Exception("Http Response (%s) Is Not 200: %s" % (url,str(resp.status_code)))

  resp_formatted = resp.text[resp.text.find('<'):len(resp.text)]
  # Write response to an XML File
  with open(file_path, 'w') as f:
    f.write(resp_formatted)

  # Load in the XML file into ElementTree
  tree = ET.parse(file_path)

  #Load into a dataframe and return the data frame
  root = tree.getroot()

  ns = {'sdmx': 'http://www.SDMX.org/resources/SDMXML/schemas/v2_0/generic'}

  df = pd.DataFrame()

  #Add column headers
  df.insert(0,date_range,[],True)
  df.insert(1,"DATE",[],True)
  index = 2
  for series in root.findall('./sdmx:DataSet/sdmx:Series',ns):
    current_country = ""
    for value in series.findall('./sdmx:SeriesKey/sdmx:Value',ns): 
      if(value.get('concept')) == 'LOCATION':
        current_country = value.get('value')
        
        df.insert(index,value.get('value'),[],True)
    index+=1

  # This needs to account for both Quarterly and Monthly data.
  date_range_list = []
  date_list = []
  year_start, qtr_start =  params['startTime'].split('-Q')
  year_end, qtr_end =  params['endTime'].split('-Q')

  match date_range:
    case 'QTR':
      #From year_start to year_end, calculate all the quarters. Populate date_range_list and date_list
      date_list = pd.date_range('%s-01-01' % (year_start),'%s-01-01' % (int(year_end)+1), freq='QS').strftime("1/%-m/%Y").tolist()
      date_ranges = pd.PeriodIndex(pd.to_datetime(date_list, format='%d/%m/%Y'),freq='Q').strftime('%Y-Q%q')

      date_range_list = date_ranges.tolist()

      # Need to align QTR and DATE between df_original and df_QoQ.
      date_list.pop(0)
      date_range_list.pop()
    case 'MTH':
      #From year_start to year_end, calculate all the months. Populate date_range_list and date_list
      date_range_list = pd.date_range('%s-01-01' % (year_start),'%s-01-01' % (int(year_end)+1), freq='MS').strftime("%Y-%m").tolist()
      date_list = pd.date_range('%s-01-01' % (year_start),'%s-01-01' % (int(year_end)+1), freq='MS').strftime("1/%-m/%Y").tolist()

  df[date_range] = date_range_list
  df['DATE'] = date_list

  #Add observations for all countries
  for series in root.findall('./sdmx:DataSet/sdmx:Series',ns):
    for value in series.findall('./sdmx:SeriesKey/sdmx:Value',ns): 
      if(value.get('concept')) == 'LOCATION':
        current_country = value.get('value')

        observations = series.findall('sdmx:Obs',ns)

        for observation in observations:
          #Get qtr, date and observation. Add to column in chronological order for all countries

          obs_qtr = observation.findall('sdmx:Time',ns)[0].text
          obs_row = df.index[df[date_range] == obs_qtr].tolist()
          obs_value = observation.findall('sdmx:ObsValue',ns)[0].get('value')

          #match based on quarter and country, then add the observation value
          df.loc[obs_row, current_country] = round(float(obs_value),9)        

  #Set the date to the correct datatype, and ensure the format accounts for the correct positioning of day and month values
  df['DATE'] = pd.to_datetime(df['DATE'],format='%d/%m/%Y')

  print("Retrieved Data for Series %s" % (dataset,))

  return df

  #except Exception as e:
    
  #  exc_type, exc_obj, exc_tb = sys.exc_info()
  #  fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
  #  print(exc_type, fname, exc_tb.tb_lineno)

def get_yf_historical_stock_data(ticker, interval, start, end):
  data = yf.download(  # or pdr.get_data_yahoo(...
    # tickers list or string as well
    tickers = ticker,

    start=start, 
    end=end, 

    # use "period" instead of start/end
    # valid periods: 1d,5d,1mo,3mo,6mo,1y,2y,5y,10y,ytd,max
    # (optional, default is '1mo')
    period = "ytd",

    # fetch data by interval (including intraday if period < 60 days)
    # valid intervals: 1m,2m,5m,15m,30m,60m,90m,1h,1d,5d,1wk,1mo,3mo
    # (optional, default is '1d')
    interval = interval,

    # group by ticker (to access via data['SPY'])
    # (optional, default is 'column')
    group_by = 'ticker',

    # adjust all OHLC automatically
    # (optional, default is False)
    auto_adjust = True,

    # download pre/post regular market hours data
    # (optional, default is False)
    prepost = True,

    # use threads for mass downloading? (True/False/Integer)
    # (optional, default is True)
    threads = True,

    # proxy URL scheme use use when downloading?
    # (optional, default is None)
    proxy = None
  )

  df_yf = data.reset_index()
  df_yf = df_yf.rename(columns={"Date": "DATE"})

  return df_yf

def get_yf_analysis(ticker):
  company = yf.Ticker(ticker)
  import pdb; pdb.set_trace()
  # get stock info
  company.info

  # get historical market data
  hist = company.history(period="max")

  # show actions (dividends, splits)
  company.actions

  # show dividends
  company.dividends

  # show splits
  company.splits

  # show financials
  company.financials
  company.quarterly_financials

  # show major holders
  company.major_holders

  # show institutional holders
  company.institutional_holders

  # show balance sheet
  company.balance_sheet
  company.quarterly_balance_sheet

  # show cashflow
  company.cashflow
  company.quarterly_cashflow

  # show earnings
  company.earnings
  company.quarterly_earnings

  # show sustainability
  company.sustainability

  # show analysts recommendations
  company.recommendations

  # show next event (earnings, etc)
  company.calendar

  # show all earnings dates
  company.earnings_dates

  # show ISIN code - *experimental*
  # ISIN = International Securities Identification Number
  company.isin

  # show options expirations
  company.options

  # show news
  company.news

  # get option chain for specific expiration
  opt = company.option_chain('YYYY-MM-DD')
  # data available via: opt.calls, opt.puts

def get_yf_key_stats(ticker):
  df_company_data = pd.DataFrame()
  url = "https://finance.yahoo.com/quote/%s/key-statistics?p=%s" % (ticker, ticker)

  page = get_page(url)

  soup = BeautifulSoup(page.content, 'html.parser')

  tables = soup.find_all('table')
  statsDict = {}

  for table in tables:
    table_rows = table.find_all('tr', recursive=True)
    emptyDict = {}

    #Get rows of data.
    for tr in table_rows:
      tds = tr.find_all('td')
      boolKey = True
      keyValueSet = False

      for td in tds:
          if boolKey:
              key = td.text.strip()
              boolKey = False                
          else:
              value = td.text.strip()
              boolKey = True
              keyValueSet = True                

          if keyValueSet:
              emptyDict[key] = value
              keyValueSet = False
    statsDict.update(emptyDict)

  df_company_data.loc[ticker, 'MARKET_CAP'] = statsDict['Market Cap (intraday)']
  df_company_data.loc[ticker, 'EV'] = statsDict['Enterprise Value']
  df_company_data.loc[ticker, 'AVG_VOL_3M'] = statsDict['Avg Vol (3 month) 3']
  df_company_data.loc[ticker, 'AVG_VOL_10D'] = statsDict['Avg Vol (10 day) 3']
  df_company_data.loc[ticker, '50_DAY_MOVING_AVG'] = statsDict['50-Day Moving Average 3']
  df_company_data.loc[ticker, '200_DAY_MOVING_AVG'] = statsDict['200-Day Moving Average 3']
  df_company_data.loc[ticker, 'EV_REVENUE'] = statsDict['Enterprise Value/Revenue']
  df_company_data.loc[ticker, 'EV_EBITDA'] = statsDict['Enterprise Value/EBITDA']
  df_company_data.loc[ticker, 'PRICE_BOOK'] = statsDict['Price/Book (mrq)']

  df_company_data = dataframe_convert_to_numeric(df_company_data, '50_DAY_MOVING_AVG')
  df_company_data = dataframe_convert_to_numeric(df_company_data, '200_DAY_MOVING_AVG')

  return df_company_data
"""
def get_yf_financials(ticker):
  df_company_data = pd.DataFrame()
  url = "https://finance.yahoo.com/quote/%s/financials/" % (ticker)

  page = get_page(url)

  soup = BeautifulSoup(page.content, 'html.parser')

  tables = soup.find_all('table')
  statsDict = {}
  import pdb; pdb.set_trace()
  for table in tables:
    table_rows = table.find_all('tr', recursive=True)
    emptyDict = {}

    #Get rows of data.
    for tr in table_rows:
      tds = tr.find_all('td')
      boolKey = True
      keyValueSet = False

      for td in tds:
          if boolKey:
              key = td.text.strip()
              boolKey = False                
          else:
              value = td.text.strip()
              boolKey = True
              keyValueSet = True                

          if keyValueSet:
              emptyDict[key] = value
              keyValueSet = False
    statsDict.update(emptyDict)

  df_company_data.loc[ticker, 'MARKET_CAP'] = statsDict['Market Cap (intraday)']
  df_company_data.loc[ticker, 'EV'] = statsDict['Enterprise Value']
  df_company_data.loc[ticker, 'AVG_VOL_3M'] = statsDict['Avg Vol (3 month) 3']
  df_company_data.loc[ticker, 'AVG_VOL_10D'] = statsDict['Avg Vol (10 day) 3']
  df_company_data.loc[ticker, '50_DAY_MOVING_AVG'] = statsDict['50-Day Moving Average 3']
  df_company_data.loc[ticker, '200_DAY_MOVING_AVG'] = statsDict['200-Day Moving Average 3']
  df_company_data.loc[ticker, 'EV_REVENUE'] = statsDict['Enterprise Value/Revenue']
  df_company_data.loc[ticker, 'EV_EBITDA'] = statsDict['Enterprise Value/EBITDA']
  df_company_data.loc[ticker, 'PRICE_BOOK'] = statsDict['Price/Book (mrq)']

  df_company_data = dataframe_convert_to_numeric(df_company_data, '50_DAY_MOVING_AVG')
  df_company_data = dataframe_convert_to_numeric(df_company_data, '200_DAY_MOVING_AVG')

  return df_company_data
"""
def get_finwiz_stock_data(ticker):
  df_company_data = pd.DataFrame()
  url_finviz = "https://finviz.com/quote.ashx?t=%s" % (ticker)
  page = get_page(url_finviz)

  soup = BeautifulSoup(page.content, 'html.parser')

  table = soup.find_all('table')
  table_rows = table[8].find_all('tr', recursive=False)

  emptyDict = {}

  #Get rows of data.
  for tr in table_rows:
      tds = tr.find_all('td')
      boolKey = True
      keyValueSet = False
      for td in tds:
          if boolKey:
              key = td.text.strip()
              boolKey = False                
          else:
              value = td.text.strip()
              boolKey = True
              keyValueSet = True                

          if keyValueSet:
              emptyDict[key] = value
              keyValueSet = False

  df_company_data.loc[ticker, 'PE'] = emptyDict['P/E']
  df_company_data.loc[ticker, 'EPS_TTM'] = emptyDict['EPS (ttm)']
  df_company_data.loc[ticker, 'PE_FORWARD'] = emptyDict['Forward P/E']
  df_company_data.loc[ticker, 'EPS_Y1'] = emptyDict['EPS next Y']
  df_company_data.loc[ticker, 'PEG'] = emptyDict['PEG']
  df_company_data.loc[ticker, 'EPS_Y0'] = emptyDict['EPS this Y']
  df_company_data.loc[ticker, 'PRICE_BOOK'] = emptyDict['P/B']
  df_company_data.loc[ticker, 'PRICE_BOOK'] = emptyDict['P/B']
  df_company_data.loc[ticker, 'PRICE_SALES'] = emptyDict['P/S']
  df_company_data.loc[ticker, 'TARGET_PRICE'] = emptyDict['Target Price']
  df_company_data.loc[ticker, 'ROE'] = emptyDict['ROE']
  df_company_data.loc[ticker, '52W_RANGE'] = emptyDict['52W Range']
  df_company_data.loc[ticker, 'QUICK_RATIO'] = emptyDict['Quick Ratio']
  df_company_data.loc[ticker, 'GROSS_MARGIN'] = emptyDict['Gross Margin']
  df_company_data.loc[ticker, 'CURRENT_RATIO'] = emptyDict['Current Ratio']

  return df_company_data

def get_stockrow_stock_data(ticker, debug):
  if debug:
    'Revenue','EBT','Net Income','PE Ratio','Earnings/Sh','Total Debt','Cash Flow/Sh','Book Value/Sh'
    data = [
        ['Revenue','156,508.00','170,910.00','182,795.00','233,715.00','215,639.00','228,572.00','265,809.00','259,968.00','274,150.00','365,817.00','393,914.00','415,323.00','431,853.00'], 
        ['EBT','55,763.00','50,155.00','53,483.00','72,515.00','61,372.00','64,089.00','72,903.00','65,737.00','67,091.00','109,207.00','119,189.00','124,010.00','127,381.00'],
        ['Net Income','41,733.00','37,037.00','39,510.00','53,394.00','45,687.00','48,351.00','59,531.00','55,256.00','57,411.00','94,680.00','99,856.00','103,545.00','106,156.00'],
        ['PE Ratio','14.76','11.95','15.56','11.93','13.59','16.63','18.99','18.76','35.12','25.09','21.20','19.90','18.90'],
        ['Earnings/Sh','1.59','1.43','1.62','2.32','2.09','2.30','2.98','2.97','3.28','5.61','6.14','6.53','6.89'],
        ['Cash Flow/Sh','1.94','2.07','2.45','3.53','3.03','3.05','3.91','3.76','4.65','6.23','7.14','7.60','8.58'],
        ['Book Value/Sh','4.52','4.77','4.58','5.19','5.86','6.42','5.41','4.90','3.77','3.78','3.69','4.10','4.82'],
        ['Total Debt','–','16,960.00','35,295.00','64,328.00','87,032.00','115,680.00','114,483.00','108,047.00','122,278.00','136,522.00','–','–','–'],
        ['EBITDA','0.00','0.00','0.00','0.00','0.00','69,428','80,342','74,542','76,395','120,233','0.00','0.00','0.00']
    ]
    df = pd.DataFrame(data, columns=['YEAR','2012','2013','2014','2015','2016','2017','2018','2019','2020','2021','2022','2023','2024'])
  else:
    page = get_page_selenium('https://stockrow.com/%s' % (ticker))

    soup = BeautifulSoup(page, 'html.parser')

    try:
      table = soup.find_all('table')[0]
    except IndexError as e:
       raise Exception("Did not load table from stockrow. Try again")

    table_rows = table.find_all('tr', recursive=True)
    table_rows_header = table.find_all('tr')[0].find_all('th')

    df = pd.DataFrame()
    index = 0

    for header in table_rows_header:
      df.insert(index,header.text,[],True)
      index+=1
    print("did we get any rows?")
    print(table_rows)
    #Get rows of data.
    for tr in table_rows:

      if(tr.find_all('td')):
        print(tr.find_all('td')[len(tr.find_all('td'))-1].text.strip())
        row_heading = tr.find_all('td')[len(tr.find_all('td'))-1].text.strip().replace("Created with Highcharts 8.2.2foo","")   
        if(row_heading in ['Revenue','EBT','Net Income','PE Ratio','Earnings/Sh','Total Debt','Cash Flow/Sh','Book Value/Sh']):
          tds = tr.find_all('td', recursive=True)
          if(tds):
            temp_row = []
            for td in tds:
              temp_row.append(td.text.strip().replace("Created with Highcharts 8.2.2foo",""))        

            df.loc[len(df.index)] = temp_row

    df.rename(columns={ df.columns[13]: "YEAR" }, inplace = True)

    # get a list of columns
    cols = list(df)

    # move the column to head of list using index, pop and insert
    cols.insert(0, cols.pop(cols.index('YEAR')))

    # reorder
    df = df.loc[:, cols]
    print("df before df2 is populated. Does it contain data?")
    print(df)
    page = get_page_selenium("https://www.wsj.com/market-data/quotes/%s/financials/annual/income-statement" % (ticker))

    soup = BeautifulSoup(page, 'html.parser')
    tables = soup.find_all('table')
    try:
      table_rows = tables[0].find_all('tr', recursive=True)
    except IndexError as e:
       raise Exception("Did not load table from wsj. Try again")

    table_rows_header = tables[0].find_all('tr')[0].find_all('th')

    #TODO: Get EBITDA from table
    
    df2 = pd.DataFrame()
    index = 0
    for header in table_rows_header:
      if(index == 0):
        df2.insert(0,"YEAR",[],True)
      else:
        df2.insert(index,header.text,[],True)
      index+=1
    
    #drop the last column has it does not contain any data but rather is a graphic of the trend
    df2 = df2.iloc[: , :-1]
    
    #Insert New Row. Format the data to show percentage as float

    for tr in table_rows:

      if(tr.find_all('td')):
        if(tr.find_all('td')[0].text in ['EBITDA']):
          temp_row = []

          td = tr.find_all('td')
          for obs in td:
            if(len(obs.text.strip()) > 0):
              text = obs.text
              temp_row.append(text)        

          df2.loc[len(df2.index)] = temp_row
          break

    df = df.append(df2,ignore_index = True)    
  #import pdb; pdb.set_trace()
  print("df after merge")
  print(df)
  #format the df

  df_transposed = transpose_df(df)

  #df_transposed = df.T
  #new_header = df_transposed.iloc[0] #grab the first row for the header
  #df_transposed = df_transposed[1:] #take the data less the header row
  #df_transposed.columns = new_header #set the header row as the df header
  
  df_transposed = df_transposed.rename(columns={
    "Revenue":"SALES",          
    "EBT":"EBIT",               
    "Net Income":"NET_INCOME",        
    "PE Ratio":"PE_RATIO",          
    "Earnings/Sh":"EARNINGS_PER_SHARE",       
    "Cash Flow/Sh":"CASH_FLOW_PER_SHARE",      
    "Book Value/Sh":"BOOK_VALUE_PER_SHARE",     
    "Total Debt":"TOTAL_DEBT",        
    "EBITDA": "EBITDA"                   
    })  

  df_transposed['EBITDA'] = df_transposed['EBITDA'].fillna("–")
  print("df_transposed before numeric conversion")
  print(df_transposed)

  #format numeric values in dataframe
  #df_transposed = df_transposed.squeeze()
  df_transposed = dataframe_convert_to_numeric(df_transposed, 'SALES')
  df_transposed = dataframe_convert_to_numeric(df_transposed, 'EBIT')
  df_transposed = dataframe_convert_to_numeric(df_transposed, 'NET_INCOME')
  df_transposed = dataframe_convert_to_numeric(df_transposed, 'PE_RATIO')
  df_transposed = dataframe_convert_to_numeric(df_transposed, 'EARNINGS_PER_SHARE')
  df_transposed = dataframe_convert_to_numeric(df_transposed, 'CASH_FLOW_PER_SHARE')
  df_transposed = dataframe_convert_to_numeric(df_transposed, 'BOOK_VALUE_PER_SHARE')
  df_transposed = dataframe_convert_to_numeric(df_transposed, 'TOTAL_DEBT')
  df_transposed = dataframe_convert_to_numeric(df_transposed, 'EBITDA')

  todays_date = date.today()
  one_year_ago = dt(todays_date.year - 1, 12, 31)
  two_year_ago = dt(todays_date.year - 2, 12, 31)
  three_year_ago = dt(todays_date.year - 3, 12, 31)
  one_year_future = dt(todays_date.year + 1, 12, 31)
  two_year_future = dt(todays_date.year + 2, 12, 31)

  list_dates = []
  list_dates.append(str(three_year_ago.year))
  list_dates.append(str(two_year_ago.year))
  list_dates.append(str(one_year_ago.year))
  list_dates.append(str(todays_date.year))
  list_dates.append(str(one_year_future.year))
  list_dates.append(str(two_year_future.year))

  df_transposed = df_transposed.loc[list_dates]

  return df_transposed

def get_zacks_balance_sheet_shares(ticker):
  df_balance_sheet_annual = pd.DataFrame()
  df_balance_sheet_quarterly = pd.DataFrame()

  #only balance sheet that shows treasury stock line item
  page = get_page('https://www.zacks.com/stock/quote/%s/balance-sheet' % (ticker))

  soup = BeautifulSoup(page.content, 'html.parser')

  table = soup.find_all('table')

  table_annual = table[4]
  table_quarterly = table[7]

  df_balance_sheet_annual = convert_html_table_to_df(table_annual,False)
  df_balance_sheet_quarterly = convert_html_table_to_df(table_quarterly,False)

  df_balance_sheet_annual = transpose_df(df_balance_sheet_annual)

  df_balance_sheet_annual = df_balance_sheet_annual.rename(columns={"Preferred Stock":"PREFERRED_STOCK"})                        
  df_balance_sheet_annual = df_balance_sheet_annual.rename(columns={"Common Stock (Par)":"COMMON_STOCK_PAR"})                          
  df_balance_sheet_annual = df_balance_sheet_annual.rename(columns={"Capital Surplus":"CAPITAL_SURPLUS"})                             
  df_balance_sheet_annual = df_balance_sheet_annual.rename(columns={"Retained Earnings":"RETAINED_EARNINGS"})                           
  df_balance_sheet_annual = df_balance_sheet_annual.rename(columns={"Other Equity":"OTHER_EQUITY"})                                
  df_balance_sheet_annual = df_balance_sheet_annual.rename(columns={"Treasury Stock":"TREASURY_STOCK"})                              
  df_balance_sheet_annual = df_balance_sheet_annual.rename(columns={"Total Shareholder's Equity":"TOTAL_SHAREHOLDERS_EQUITY"})                  
  df_balance_sheet_annual = df_balance_sheet_annual.rename(columns={"Total Liabilities & Shareholder's Equity":"TOTAL_LIABILITIES_SHAREHOLDERS_EQUITY"})    
  df_balance_sheet_annual = df_balance_sheet_annual.rename(columns={"Total Common Equity":"TOTAL_COMMON_EQUITY"})                         
  df_balance_sheet_annual = df_balance_sheet_annual.rename(columns={"Shares Outstanding":"SHARES_OUTSTANDING"})                          
  df_balance_sheet_annual = df_balance_sheet_annual.rename(columns={"Book Value Per Share":"BOOK_VALUE_PER_SHARE"})  
  df_balance_sheet_annual = dataframe_convert_to_numeric(df_balance_sheet_annual,'PREFERRED_STOCK')
  df_balance_sheet_annual = dataframe_convert_to_numeric(df_balance_sheet_annual,'COMMON_STOCK_PAR')
  df_balance_sheet_annual = dataframe_convert_to_numeric(df_balance_sheet_annual,'CAPITAL_SURPLUS')
  df_balance_sheet_annual = dataframe_convert_to_numeric(df_balance_sheet_annual,'RETAINED_EARNINGS')
  df_balance_sheet_annual = dataframe_convert_to_numeric(df_balance_sheet_annual,'OTHER_EQUITY')
  df_balance_sheet_annual = dataframe_convert_to_numeric(df_balance_sheet_annual,'TREASURY_STOCK')
  df_balance_sheet_annual = dataframe_convert_to_numeric(df_balance_sheet_annual,'TOTAL_SHAREHOLDERS_EQUITY')
  df_balance_sheet_annual = dataframe_convert_to_numeric(df_balance_sheet_annual,'TOTAL_LIABILITIES_SHAREHOLDERS_EQUITY')
  df_balance_sheet_annual = dataframe_convert_to_numeric(df_balance_sheet_annual,'TOTAL_COMMON_EQUITY')
  df_balance_sheet_annual = dataframe_convert_to_numeric(df_balance_sheet_annual,'SHARES_OUTSTANDING')
  df_balance_sheet_annual = dataframe_convert_to_numeric(df_balance_sheet_annual,'BOOK_VALUE_PER_SHARE')
  df_balance_sheet_annual.reset_index(inplace=True)
  df_balance_sheet_annual = df_balance_sheet_annual.rename(columns = {'index':'DATE'})
  df_balance_sheet_annual['DATE'] = pd.to_datetime(df_balance_sheet_annual['DATE'],format='%m/%d/%Y')
  df_balance_sheet_annual = df_balance_sheet_annual.rename_axis(None, axis=1)

  df_balance_sheet_quarterly = transpose_df(df_balance_sheet_quarterly)
  df_balance_sheet_quarterly = df_balance_sheet_quarterly.rename(columns={"Preferred Stock":"PREFERRED_STOCK"})                        
  df_balance_sheet_quarterly = df_balance_sheet_quarterly.rename(columns={"Common Stock (Par)":"COMMON_STOCK_PAR"})                          
  df_balance_sheet_quarterly = df_balance_sheet_quarterly.rename(columns={"Capital Surplus":"CAPITAL_SURPLUS"})                             
  df_balance_sheet_quarterly = df_balance_sheet_quarterly.rename(columns={"Retained Earnings":"RETAINED_EARNINGS"})                           
  df_balance_sheet_quarterly = df_balance_sheet_quarterly.rename(columns={"Other Equity":"OTHER_EQUITY"})                                
  df_balance_sheet_quarterly = df_balance_sheet_quarterly.rename(columns={"Treasury Stock":"TREASURY_STOCK"})                              
  df_balance_sheet_quarterly = df_balance_sheet_quarterly.rename(columns={"Total Shareholder's Equity":"TOTAL_SHAREHOLDERS_EQUITY"})                  
  df_balance_sheet_quarterly = df_balance_sheet_quarterly.rename(columns={"Total Liabilities & Shareholder's Equity":"TOTAL_LIABILITIES_SHAREHOLDERS_EQUITY"})    
  df_balance_sheet_quarterly = df_balance_sheet_quarterly.rename(columns={"Total Common Equity":"TOTAL_COMMON_EQUITY"})                         
  df_balance_sheet_quarterly = df_balance_sheet_quarterly.rename(columns={"Shares Outstanding":"SHARES_OUTSTANDING"})                          
  df_balance_sheet_quarterly = df_balance_sheet_quarterly.rename(columns={"Book Value Per Share":"BOOK_VALUE_PER_SHARE"})  
  df_balance_sheet_quarterly = dataframe_convert_to_numeric(df_balance_sheet_quarterly,'PREFERRED_STOCK')
  df_balance_sheet_quarterly = dataframe_convert_to_numeric(df_balance_sheet_quarterly,'COMMON_STOCK_PAR')
  df_balance_sheet_quarterly = dataframe_convert_to_numeric(df_balance_sheet_quarterly,'CAPITAL_SURPLUS')
  df_balance_sheet_quarterly = dataframe_convert_to_numeric(df_balance_sheet_quarterly,'RETAINED_EARNINGS')
  df_balance_sheet_quarterly = dataframe_convert_to_numeric(df_balance_sheet_quarterly,'OTHER_EQUITY')
  df_balance_sheet_quarterly = dataframe_convert_to_numeric(df_balance_sheet_quarterly,'TREASURY_STOCK')
  df_balance_sheet_quarterly = dataframe_convert_to_numeric(df_balance_sheet_quarterly,'TOTAL_SHAREHOLDERS_EQUITY')
  df_balance_sheet_quarterly = dataframe_convert_to_numeric(df_balance_sheet_quarterly,'TOTAL_LIABILITIES_SHAREHOLDERS_EQUITY')
  df_balance_sheet_quarterly = dataframe_convert_to_numeric(df_balance_sheet_quarterly,'TOTAL_COMMON_EQUITY')
  df_balance_sheet_quarterly = dataframe_convert_to_numeric(df_balance_sheet_quarterly,'SHARES_OUTSTANDING')
  df_balance_sheet_quarterly = dataframe_convert_to_numeric(df_balance_sheet_quarterly,'BOOK_VALUE_PER_SHARE')
  df_balance_sheet_quarterly.reset_index(inplace=True)
  df_balance_sheet_quarterly = df_balance_sheet_quarterly.rename(columns = {'index':'DATE'})
  df_balance_sheet_quarterly['DATE'] = pd.to_datetime(df_balance_sheet_quarterly['DATE'],format='%m/%d/%Y')
  df_balance_sheet_quarterly = df_balance_sheet_quarterly.rename_axis(None, axis=1)

  return df_balance_sheet_annual, df_balance_sheet_quarterly

def get_zacks_peer_comparison(ticker):
  df_peer_comparison = pd.DataFrame()
  url = "https://www.zacks.com/stock/research/%s/industry-comparison" % (ticker)

  page = get_page(url)
  soup = BeautifulSoup(page.content, 'html.parser')
  table = soup.find_all('table')

  table_peer_comparison = table[2]

  df_peer_comparison = convert_html_table_to_df(table_peer_comparison,True)
  new = df_peer_comparison["Symbol"].str.split(' ',n=1,expand=True)
  df_peer_comparison["Ticker"] = new[0]

  df_peer_comparison = df_peer_comparison.drop(['Zacks Rank', 'Symbol'], axis=1)

  return df_peer_comparison

def get_zacks_earnings_surprises(ticker):
  df_earnings_release_date = pd.DataFrame()
  df_earnings_surprises = pd.DataFrame()
  df_sales_surprises = pd.DataFrame()

  url = "https://www.zacks.com/stock/research/%s/earnings-calendar" % (ticker)

  page = get_page(url)
  soup = BeautifulSoup(page.content, 'html.parser')

  #Get Earnings Release Date
  table_earnings_release_date = soup.find_all('table')[2]
  df_earnings_release_date = convert_html_table_to_df(table_earnings_release_date,True)
  df_earnings_release_date['Release Date'] = df_earnings_release_date['Report Date'].str[:10]
  df_earnings_release_date = df_earnings_release_date.drop(['Zacks Consensus Estimate', 'Earnings ESP','Report Date'], axis=1)
  df_earnings_release_date['Release Date'] = pd.to_datetime(df_earnings_release_date['Release Date'],format='%m/%d/%Y')

  #Need to extract Earnings and Sales Surprises data from json object in javascript on page
  scripts = soup.find_all('script')[29]
  match_pattern = re.compile(r'(?<=\= ).*\}')
  match_string = scripts.text.strip().replace('\n','')
  matches = match_pattern.findall(match_string)
  match_string = matches[0]

  json_object = json.loads(match_string)

  list_earnings_announcements_earnings = json_object['earnings_announcements_earnings_table']
  list_earnings_announcements_sales = json_object['earnings_announcements_sales_table']

  df_earnings_surprises = convert_list_to_df(list_earnings_announcements_earnings)
  df_earnings_surprises = df_earnings_surprises.drop(df_earnings_surprises.iloc[:, 4:7],axis = 1)
  df_earnings_surprises.rename(columns={ df_earnings_surprises.columns[0]: "Date",df_earnings_surprises.columns[1]: "Period",df_earnings_surprises.columns[2]: "Estimate",df_earnings_surprises.columns[3]: "Reported" }, inplace = True)
  df_earnings_surprises['Date'] = pd.to_datetime(df_earnings_surprises['Date'],format='%m/%d/%y')
  df_earnings_surprises = dataframe_convert_to_numeric(df_earnings_surprises,'Estimate')
  df_earnings_surprises = dataframe_convert_to_numeric(df_earnings_surprises,'Reported')

  df_sales_surprises = convert_list_to_df(list_earnings_announcements_sales)
  df_sales_surprises = df_sales_surprises.drop(df_sales_surprises.iloc[:, 4:7],axis = 1)
  df_sales_surprises.rename(columns={ df_sales_surprises.columns[0]: "Date",df_sales_surprises.columns[1]: "Period",df_sales_surprises.columns[2]: "Estimate",df_sales_surprises.columns[3]: "Reported" }, inplace = True)
  df_sales_surprises['Date'] = pd.to_datetime(df_sales_surprises['Date'],format='%m/%d/%y')
  df_sales_surprises = dataframe_convert_to_numeric(df_sales_surprises,'Estimate')
  df_sales_surprises = dataframe_convert_to_numeric(df_sales_surprises,'Reported')

  return df_earnings_release_date, df_earnings_surprises, df_sales_surprises

def get_zacks_product_line_geography(ticker):
  df_product_line = pd.DataFrame()
  df_geography = pd.DataFrame()
  pd.set_option('display.max_colwidth', None)

  url = "https://www.zacks.com/stock/research/%s/key-company-metrics-details" % (ticker)

  page = get_page(url)
  soup = BeautifulSoup(page.content, 'html.parser')
  table = soup.find_all('table')

  table_product_line_geography = soup.find_all('table')[2]
  table_rows = table_product_line_geography.find_all('tr')
  
  file_dict = {}
  df = pd.DataFrame()

  #Insert New Row. Format the data to show percentage as float
  for tr in table_rows:
    temp_row = []
    table_rows_header = tr.find_all('th')

    if(len(table_rows_header) > 0):
      if(df.shape[0] > 0):
        file_dict[header_text] = copy.copy(df)     
        df = pd.DataFrame()
      index = 0

      for header in table_rows_header:
        df.insert(index,str(header.text).strip(),[],True)
        index+=1
      header_text = table_rows_header[0].text
      table_rows_header = []
    else:
      td = tr.find_all('td')
      for obs in td:      
        if(obs.p):
          text = obs.p.attrs['title']
        else:
          text = str(obs.text).strip()
        temp_row.append(text)        

      if(len(temp_row) == len(df.columns)):
        df.loc[len(df.index)] = temp_row

  df_product_line = file_dict['Revenue - Line of Business Segments']
  df_geography = file_dict['Revenue - Geographic Segments']

  # Clean up dataframes
  df_product_line = df_product_line.drop(columns='YR Estimate', axis=1)
  df_product_line = df_product_line.iloc[:, 0:2]
  colname = df_product_line.columns[1]
  df_product_line = dataframe_convert_to_numeric(df_product_line,colname)
  df_product_line = df_product_line.iloc[:4,:]

  df_geography = df_geography.drop(columns='YR Estimate', axis=1)
  df_geography = df_geography.iloc[:, 0:2]
  colname = df_geography.columns[1]
  df_geography = dataframe_convert_to_numeric(df_geography,colname)
  df_geography = df_geography.iloc[:4,:]

  return df_product_line, df_geography

def get_api_json_data(url, filename):

    #check if current file has todays system date, and if it does load from current file. Otherwise, continue to call the api
    file_path = "%s/JSON/%s" % (sys.path[0],filename)
    data_list = []

    todays_date = date.today()
    try:
      file_mod_date = time.ctime(os.path.getmtime(file_path))
      file_mod_date = dt.strptime(file_mod_date, '%a %b %d %H:%M:%S %Y')
    except FileNotFoundError as e:
      #Set file mod date to nothing as we do not have a file
      file_mod_date = None

    try:
        #Check if file date is today. If so, continue. Otherwise, throw exception so that we can use the API instead to load the data
        if(file_mod_date.date() == todays_date):
            my_file = open(file_path, "r")        
        else:
            # Throw exception so that we can read the data from api
            raise Exception('Need to read from API') 
    except Exception as error:
        temp_data = []
        temp_data.append(requests.get(url).json())

        # Write response to File
        with open(file_path, 'w') as f:
            for item in temp_data:
                f.write("%s\n" % item)

        # try to open the file in read mode again
        my_file = open(file_path, "r")        

    data = my_file.read()
    
    # replacing end splitting the text 
    # when newline ('\n') is seen.
    liststr = data.split("\n")
    my_file.close()

    data_list = eval(liststr[0])

    return data_list

def get_api_json_data_no_file(url):

    data_list = []

    data_list.append(requests.get(url).json())

    return data_list


# Get S&P500 Monthly Close Prices from YF #
def get_sp500_monthly_prices():
  todays_date = date.today()
  date_str = "%s-%s-%s" % (todays_date.year, todays_date.month, "01")

  # Get S&P500 close month intervals using above date range
  df_SP500 = get_yf_historical_stock_data("^GSPC", "1mo", "1959-01-01", date_str)

  #Remove unnecessary columns from df_SP500 and rename columns
  df_SP500 = df_SP500.drop(['Open', 'High', 'Low', 'Volume'], axis=1)
  df_SP500 = df_SP500.rename(columns={"Close": "SP500"})

  return df_SP500

def get_invest_data(country_list, bond_year, from_date):

  todays_date = date.today()
  todays_date_full = "%s/%s/%s" % (todays_date.day, todays_date.month, todays_date.year)

  df = pd.DataFrame()
  df.insert(0,"DATE",[],True)

  for country in country_list:
    bond = "%s %sy" % (country, bond_year)

    print("Getting %s-Y Data For: %s" % (bond_year, country,))

    time.sleep(10)
    try:
      data = invest.get_bond_historical_data(bond=bond, from_date=from_date, to_date=todays_date_full)

      # Transformations to do: 
      # 1) get the relevant column
      # 2) convert series to df
      # 3) make the index a column
      # 4) Rename columns
      # 5) Append to master df

      data = data['Close']
      data = data.to_frame()
      data.reset_index(level=0, inplace=True)
      data = data.rename(columns={"Date": "DATE" ,"Close": country})
      df = append_two_df(df, data)

    except RuntimeError as e:
      print("NO %s-Y DATA FOR: %s" % (bond_year, country,))

  df = df.sort_values(by='DATE', ignore_index=True)

  return df

def scrape_world_gdp_table(url):
  #Scrape GDP Table from Trading Economics
  #url = "https://tradingeconomics.com/matrix"

  page = get_page(url)

  soup = BeautifulSoup(page.content, 'html.parser')
  table = soup.find('table')
  #table_rows = table.find_all('tr', attrs={'align':'center'})
  table_rows = table.find_all('tr')
  table_rows_header = table.find_all('tr')[0].find_all('th')
  df = pd.DataFrame()
  index = 0
  for header in table_rows_header:
    if(index == 0):
      df.insert(0,"Country",[],True)
    else:
      df.insert(index,header.text,[],True)

    index+=1

  #Insert New Row. Format the data to show percentage as float

  for tr in table_rows:
    temp_row = []
    first_col = True

    td = tr.find_all('td')
    if len(td) > 0:
      for obs in td:
        if(first_col):
          text = ''.join(e for e in obs.text if e.isalnum())
          text = re.sub("([A-Z])", " \\1", text).strip()
        else:
          if(obs.text.find('%') < 0):
            text = obs.text
          else:
            text = obs.text.strip('%')
            text = float(text.strip('%'))/100
        temp_row.append(text)        
        first_col = False

      df.loc[len(df.index)] = temp_row

  return df

def get_ism_manufacturing_content():

  ism_date, ism_month = get_ism_date(1)
  url_ism = get_ism_manufacturing_url(ism_month)

  #This is duplicate code found in get_page function but we need to handle special case of ism data where page may not be found and we need to switch to 1 month previous
  header={'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2227.0 Safari/537.36'}
  page = requests.get(url=url_ism,headers=header)

  try:
      page.raise_for_status()
  except requests.exceptions.HTTPError as e:
    if(page.status_code == 404):
        # Use previous month to get ISM data
        ism_date, ism_month = get_ism_date(2)
        url_ism = get_ism_manufacturing_url(ism_month)

        page = get_page(url_ism)

    else:
        raise Exception("Http Response (%s) Is Not 200: %s" % (url_ism,str(page.status_code)))

  return ism_date, ism_month, page

def get_ism_services_content():

  ism_date, ism_month = get_ism_date(1)
  url_ism = get_ism_services_url(ism_month)

  #This is duplicate code found in get_page function but we need to handle special case of ism data where page may not be found and we need to switch to 1 month previous
  header={'User-Agent':'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2227.0 Safari/537.36'}
  page = requests.get(url=url_ism,headers=header)

  try:
      page.raise_for_status()
  except requests.exceptions.HTTPError as e:
    if(page.status_code == 404):
        # Use previous month to get ISM data
        ism_date, ism_month = get_ism_date(2)
        url_ism = get_ism_services_url(ism_month)

        page = get_page(url_ism)

    else:
        raise Exception("Http Response (%s) Is Not 200: %s" % (url_ism,str(page.status_code)))

  return ism_date, ism_month, page


def get_ism_manufacturing_url(ism_month):
  url_ism = 'https://www.ismworld.org/supply-management-news-and-reports/reports/ism-report-on-business/pmi/%s' % (ism_month.lower(),)

  return url_ism

def get_ism_services_url(ism_month):
  url_ism = 'https://www.ismworld.org/supply-management-news-and-reports/reports/ism-report-on-business/services/%s' % (ism_month.lower(),)

  return url_ism

def get_ism_date(delta):
  #get date range
  todays_date = date.today()

  #use todays date to get ism month (first day of last month) and use the date in scraping functions
  ism_date = todays_date - relativedelta.relativedelta(months=delta)
  ism_date = "01-%s-%s" % (ism_date.month, ism_date.year) #make the ism date the first day of ism month
  ism_date = dt.strptime(ism_date, "%d-%m-%Y")
  ism_month = ism_date.strftime("%B")

  return ism_date, ism_month


def scrape_ism_manufacturing_headline_index(ism_date, ism_month):

  url_ism = get_ism_manufacturing_url(ism_month)

  page = get_page(url_ism)

  soup = BeautifulSoup(page.content, 'html.parser')

  #Get all html tables on the page
  tables = soup.find_all('table')    
  table_at_a_glance = tables[0]
  
  #Convert the tables into dataframes so that we can read the data
  df_at_a_glance = convert_html_table_to_df(table_at_a_glance, True)

  #Drop Unnecessary Columns
  column_numbers = [x for x in range(df_at_a_glance.shape[1])]  # list of columns' integer indices
  column_numbers .remove(2) #removing column integer index 0
  column_numbers .remove(3)
  column_numbers .remove(4)
  column_numbers .remove(5)
  column_numbers .remove(6)
  df_at_a_glance = df_at_a_glance.iloc[:, column_numbers] #return all columns except the 0th column

  #Flip df around
  df_at_a_glance = df_at_a_glance.T

  #Rename Columns
  df_at_a_glance = df_at_a_glance.rename(columns={0: "ISM", 1:"NEW_ORDERS",2:"PRODUCTION",3:"EMPLOYMENT",4:"DELIVERIES",
                                                  5:"INVENTORIES",6:"CUSTOMERS_INVENTORIES",7:"PRICES",8:"BACKLOG_OF_ORDERS",9:"EXPORTS",10:"IMPORTS"})

  #Drop the first row because it contains the old column names
  df_at_a_glance = df_at_a_glance.iloc[1: , :]
  df_at_a_glance = df_at_a_glance.reset_index()
  df_at_a_glance = df_at_a_glance.drop(columns='index', axis=1)

  #Fix datatypes of df_at_a_glance
  for column in df_at_a_glance:
      df_at_a_glance[column] = pd.to_numeric(df_at_a_glance[column])

  #Add DATE column to df
  df_at_a_glance["DATE"] = [ism_date]

  # Reorder Columns
  # get a list of columns
  cols = list(df_at_a_glance)
  cols.insert(0, cols.pop(cols.index('DATE')))
  cols.insert(1, cols.pop(cols.index('NEW_ORDERS')))
  cols.insert(2, cols.pop(cols.index('IMPORTS')))
  cols.insert(3, cols.pop(cols.index('BACKLOG_OF_ORDERS')))
  cols.insert(4, cols.pop(cols.index('PRICES')))
  cols.insert(5, cols.pop(cols.index('PRODUCTION')))
  cols.insert(6, cols.pop(cols.index('CUSTOMERS_INVENTORIES')))
  cols.insert(7, cols.pop(cols.index('INVENTORIES')))
  cols.insert(8, cols.pop(cols.index('DELIVERIES')))
  cols.insert(9, cols.pop(cols.index('EMPLOYMENT')))
  cols.insert(10, cols.pop(cols.index('EXPORTS')))
  cols.insert(11, cols.pop(cols.index('ISM')))

  # reorder
  df_at_a_glance = df_at_a_glance[cols]

  return df_at_a_glance

def scrape_ism_services_headline_index(ism_date, ism_month):

    url_ism = get_ism_services_url(ism_month)

    page = get_page(url_ism)

    soup = BeautifulSoup(page.content, 'html.parser')

    #Get all html tables on the page
    tables = soup.find_all('table')    
    table_at_a_glance = tables[0]
    
    #Convert the tables into dataframes so that we can read the data
    #df_at_a_glance = convert_html_table_to_df(table_at_a_glance, True)

    table_rows = table_at_a_glance.find_all('tbody')[0].find_all('tr')
    table_rows_header = table_at_a_glance.find_all('tr')[1].find_all('th')
    df_at_a_glance = pd.DataFrame()

    index = 0

    for header in table_rows_header:
        df_at_a_glance.insert(index,str(header.text).strip(),[],True)
        index+=1

    #Insert New Row. Format the data to show percentage as float
    for tr in table_rows:
        temp_row = []

        tr_th = tr.find('th')
        text = str(tr_th.text).strip()
        temp_row.append(text)        

        td = tr.find_all('td')
        for obs in td:
            text = str(obs.text).strip()
            temp_row.append(text)        
        
        if(len(temp_row) == len(df_at_a_glance.columns)):
            df_at_a_glance.loc[len(df_at_a_glance.index)] = temp_row
    
    #Drop Unnecessary Columns
    column_numbers = [x for x in range(df_at_a_glance.shape[1])]  # list of columns' integer indices
    column_numbers .remove(7)
    column_numbers .remove(8)
    column_numbers .remove(9)

    df_at_a_glance = df_at_a_glance.iloc[:, column_numbers] #return all columns except the 0th column

    #Flip df around
    df_at_a_glance = df_at_a_glance.T

    # Rename Columns as per requirements of excel file 017
    df_at_a_glance = df_at_a_glance.rename(columns={0: "ISM_SERVICES", 1:"BUSINESS_ACTIVITY",2:"NEW_ORDERS",3:"EMPLOYMENT",4:"DELIVERIES",
                                                    5:"INVENTORIES",6:"PRICES",7:"BACKLOG_OF_ORDERS",8:"EXPORTS",9:"IMPORTS",10:"INVENTORY_SENTIMENT",11:"CUSTOMER_INVENTORIES"})

    #Drop the first row because it contains the old column names
    df_at_a_glance = df_at_a_glance.iloc[1: , :]
    df_at_a_glance = df_at_a_glance.head(1)
    df_at_a_glance = df_at_a_glance.reset_index()
    df_at_a_glance = df_at_a_glance.drop(columns='index', axis=1)
    df_at_a_glance = df_at_a_glance.drop(columns='CUSTOMER_INVENTORIES', axis=1)

    #Fix datatypes of df_at_a_glance
    for column in df_at_a_glance:
        df_at_a_glance[column] = pd.to_numeric(df_at_a_glance[column])

    #Add DATE column to df
    df_at_a_glance["DATE"] = [ism_date]

    # Put DATE as the first column
    # get a list of columns
    cols = list(df_at_a_glance)
    cols.insert(0, cols.pop(cols.index('DATE')))

    # reorder
    df_at_a_glance = df_at_a_glance[cols]

    return df_at_a_glance


####################
# Output Functions #
####################

def convert_excelsheet_to_dataframe(excel_file_path,sheet_name,date_exists=False, index_col=None, date_format='%d/%m/%Y'):

  filepath = os.path.realpath(__file__)
  excel_file_path = filepath[:filepath.rfind('/')] + excel_file_path

  if(index_col):
    df = pd.read_excel(excel_file_path, sheet_name=sheet_name, index_col=index_col, engine='openpyxl')
  else:
    df = pd.read_excel(excel_file_path, sheet_name=sheet_name, engine='openpyxl')

  if(date_exists):
    df['DATE'] = pd.to_datetime(df['DATE'],format=date_format)

  return df

def write_to_directory(df,filename):
    #Write to a csv file in the correct directory
    userhome = os.path.expanduser('~')
    file_name = os.path.join(userhome, 'Desktop', 'Trading_Excel_Files', 'Database',filename)
    df.to_csv(file_name, index=False)

def write_dataframe_to_excel(excel_file_path,sheet_name, df, include_index, date_position=None, format_header=False):

  filepath = os.path.realpath(__file__)
  excel_file_path = filepath[:filepath.rfind('/')] + excel_file_path

  book = openpyxl.load_workbook(excel_file_path, read_only=False, keep_vba=True)
  sheet = book[sheet_name]

  book.active = sheet

  # Delete all rows after the header so that we can replace them with our df  
  sheet.delete_rows(1,sheet.max_row)
  
  if(include_index):
    df.reset_index(level=0, inplace=True)
    # Need to remove additional unnecessary rows from beginning of df_QoQ dataframe
    df = df.iloc[1: , :].reset_index(drop=True)    

  #Write values from the df to the sheet
  for r in dataframe_to_rows(df,index=False, header=True):
    sheet.append(r)

  if(date_position >= 0):
    for row in sheet[2:sheet.max_row]: # skip the header
      cell = row[date_position]   # column date_position is a Date Field.
      cell.number_format = 'dd-mm-YYYY'

  if(format_header):  
    new_fill  = PatternFill(start_color='bde9f7', end_color='bde9f7', fill_type='solid')
    for cell in sheet[1:1]:
      cell.font = Font(bold=True)
      cell.fill = new_fill

  book.save(excel_file_path)
  book.close()

def download_file(url, save_path, chunk_size=128):
    filepath = os.path.realpath(__file__)
    excel_file_path = filepath[:filepath.rfind('/')] + save_path

    r = requests.get(url, stream=True)
    with open(excel_file_path, 'wb') as fd:
        for chunk in r.iter_content(chunk_size=chunk_size):
            fd.write(chunk)

def unzip_file(zip_file_path, zip_file, chunk_size=128):

    filepath = os.path.realpath(__file__)
    zip_file_path = filepath[:filepath.rfind('/')] + zip_file_path
    zip_file = filepath[:filepath.rfind('/')] + zip_file

    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
      zip_ref.extractall(zip_file_path)

    os.remove(zip_file)

    print("The file has been deleted successfully")

def write_value_to_cell_excel(excel_file_path,sheet_name, row, column, value):
  filepath = os.path.realpath(__file__)
  excel_file_path = filepath[:filepath.rfind('/')] + excel_file_path

  book = openpyxl.load_workbook(excel_file_path, read_only=False, keep_vba=True)
  sheet = book[sheet_name]
  book.active = sheet

  #sheet['A1'] = 1
  sheet.cell(row=row, column=column).value = value

  book.save(excel_file_path)
  book.close()

def check_sheet_exists(excel_file_path,sheet_name):
  filepath = os.path.realpath(__file__)
  excel_file_path = filepath[:filepath.rfind('/')] + excel_file_path

  book = openpyxl.load_workbook(excel_file_path, read_only=False, keep_vba=True)
      
  if sheet_name in book.sheetnames:
      return True

  return False

def create_sheet(excel_file_path, sheet, sheet_copy_name=False):
  filepath = os.path.realpath(__file__)
  excel_file_path = filepath[:filepath.rfind('/')] + excel_file_path

  book = openpyxl.load_workbook(excel_file_path, read_only=False, keep_vba=True)

  if(sheet_copy_name):
    book.active = book[sheet_copy_name]
    source = book.active
    target = book.copy_worksheet(source)
    target.title = sheet
  else:
    book.create_sheet(sheet)
    book.active = book[sheet]
    target = book.active

  target.sheet_properties.tabColor = "1072BA"
  target.sheet_view.zoomScale = 160
  book.save(excel_file_path)
  book.close()

####################
# Helper Functions #
####################


def combine_df(df_original, df_new):

  return df_original.combine(df_new, take_larger, overwrite=False)  

def append_two_df(df1, df2):
  merged_data = pd.merge(df1, df2, how='outer', on='DATE')
  return merged_data

def take_larger(s1, s2):
  return s2

def combine_df_on_index(df1, df2, index_col):
  df1 = df1.set_index(index_col)
  df2 = df2.set_index(index_col)

  return df1.combine_first(df2).reset_index()

def convert_html_table_to_df(table, contains_th):
  
  table_rows = table.find_all('tr')
  table_rows_header = table.find_all('tr')[0].find_all('th')
  df = pd.DataFrame()

  index = 0

  for header in table_rows_header:
    df.insert(index,str(header.text).strip(),[],True)
    index+=1

  #Insert New Row. Format the data to show percentage as float
  for tr in table_rows:
    temp_row = []

    if(contains_th):
      tr_th = tr.find('th')
      text = str(tr_th.text).strip()
      temp_row.append(text)        

    td = tr.find_all('td')
    for obs in td:
      
      exclude = False

      if(obs.find_all('div')):
        if 'hidden' in obs.find_all('div')[0].attrs['class']:
          exclude = True

      if not exclude:
        text = str(obs.text).strip()
        temp_row.append(text)        

    if(len(temp_row) == len(df.columns)):
      df.loc[len(df.index)] = temp_row
  
  return df

def convert_list_to_df(list):

  df = pd.DataFrame(list)

  return df

#Util function for transforming ISM data on sheet 016
def _transform_data(excel_file_path, sheet_name_original, sheet_name_new):

  df_original = convert_excelsheet_to_dataframe(excel_file_path, sheet_name_original, False)

  df_original_transposed = df_original.transpose()

  new_header = df_original_transposed.iloc[0] #grab the first row for the header
  df1 = df_original_transposed[1:] #take the data less the header row
  df1.columns = new_header #set the header row as the df header
  df1 = df1.reset_index()
  df1 = df1.rename(columns={"index": "DATE"})

  for column in df1:
    if(column != 'DATE'):
      df1[column] = pd.to_numeric(df1[column])

  # Write the updated df back to the excel sheet
  write_dataframe_to_excel(excel_file_path, sheet_name_new, df1, False, 0)

def _util_check_diff_list(li1, li2):
  # Python code to get difference of two lists
  return list(set(li1) - set(li2))

def dataframe_convert_to_numeric(df, column):
  #TODO: Deal with percentages and negative values in brackets
  try:
    df[column] = df[column].str.replace('NA','')
    df[column] = df[column].str.replace('$','', regex=False)
    df[column] = df[column].str.replace('--','')
    df[column] = df[column].str.replace(',','').replace('–','0.00')
    df[column] = df[column].str.replace('(','-', regex=True)
    df[column] = df[column].str.replace(')','', regex=True)
  except KeyError as e:
    print(df)
    print(column)

  df[column] = pd.to_numeric(df[column])

  return df

def transpose_df(df):
  df = df.T
  new_header = df.iloc[0] #grab the first row for the header
  df = df[1:] #take the data less the header row
  df.columns = new_header #set the header row as the df header

  return df

